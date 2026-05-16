"""
Parse NTHU full-university Excel course files and output all_courses.json.
Usage: python parse_excel_courses.py
"""
import io
import json
import re
import sys

import openpyxl

SHANG_PATH = r'D:\桌面\NTHU\DBMS_coursefiles\114_all_courses\11410_course_data.xlsx'
XIA_PATH   = r'D:\桌面\NTHU\DBMS_coursefiles\114_all_courses\11420_course_data.xlsx'
OUTPUT     = r'd:\桌面\NTHU\DBMS_coursefiles\course_review_system\all_courses.json'

PREFIX_RE   = re.compile(r'^\d{5}([A-Za-z]+)')
CJK_LEAD_RE = re.compile(r'^([⺀-⿿㐀-鿿豈-﫿]+)')


def normalize_id(raw) -> str:
    return re.sub(r'\s+', '', str(raw).strip()) if raw else ''


def extract_prefix(course_id: str) -> str:
    m = PREFIX_RE.match(course_id)
    return m.group(1) if m else ''


def first_prof_comma(raw) -> str:
    """114上：逗號分隔，取第一位並去空白"""
    if not raw:
        return '未知'
    name = str(raw).split(',')[0].strip()
    return name or '未知'


def first_prof_cjk(raw) -> str:
    """114下：取開頭連續中文字（名字），忽略後面的英文羅馬拼音"""
    if not raw:
        return '未知'
    text = str(raw).strip()
    m = CJK_LEAD_RE.match(text)
    if m:
        return m.group(1)
    # fallback：取第一個空白前的部分
    return text.split()[0] if text else '未知'


def parse_114_shang(path: str):
    """解析 114上 Excel，同時建立 prefix_code → 系所全名 對應表"""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    courses: list[dict] = []
    prefix_dept_map: dict[str, str] = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:          # 跳過表頭
            continue
        course_id_raw = row[0]
        if not course_id_raw:
            continue

        course_id = normalize_id(course_id_raw)
        if not course_id:
            continue

        course_name = str(row[1]).strip() if row[1] else ''
        if not course_name:
            continue

        try:
            credits = int(str(row[3]).strip())
        except (ValueError, TypeError):
            continue

        professor_name = first_prof_comma(row[6])

        department = str(row[17]).strip() if row[17] else ''
        if not department:
            department = '其他'

        # 用 代碼欄(col 16) 累積 prefix → 系所全名 對應
        code = str(row[16]).strip() if row[16] else ''
        if code and department != '其他':
            prefix_dept_map[code] = department

        courses.append({
            'course_id':      course_id,
            'course_name':    course_name,
            'credits':        credits,
            'professor_name': professor_name,
            'department':     department,
            'semester':       '114-1',
        })

    wb.close()
    return courses, prefix_dept_map


def parse_114_xia(path: str, prefix_dept_map: dict[str, str]):
    """解析 114下 Excel，用 prefix_dept_map 推斷系所"""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    courses: list[dict] = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        course_id_raw = row[0]
        if not course_id_raw:
            continue

        course_id = normalize_id(course_id_raw)
        if not course_id:
            continue

        course_name = str(row[1]).strip() if row[1] else ''
        if not course_name:
            continue

        try:
            credits = int(str(row[3]).strip())
        except (ValueError, TypeError):
            continue

        professor_name = first_prof_cjk(row[12])

        prefix = extract_prefix(course_id)
        department = prefix_dept_map.get(prefix, '其他')

        courses.append({
            'course_id':      course_id,
            'course_name':    course_name,
            'credits':        credits,
            'professor_name': professor_name,
            'department':     department,
            'semester':       '114-2',
        })

    wb.close()
    return courses


def main():
    print('【步驟 1】解析 114上 Excel ...', flush=True)
    courses_shang, prefix_dept_map = parse_114_shang(SHANG_PATH)
    print(f'  -> 解析 {len(courses_shang)} 筆，建立 {len(prefix_dept_map)} 個系所對應')

    print('【步驟 1】解析 114下 Excel ...', flush=True)
    courses_xia = parse_114_xia(XIA_PATH, prefix_dept_map)
    print(f'  -> 解析 {len(courses_xia)} 筆')

    print('\n【步驟 2】合併並去重（course_id 重複取第一筆）...', flush=True)
    all_courses: dict[str, dict] = {}
    dup_count = 0
    for c in courses_shang + courses_xia:
        if c['course_id'] not in all_courses:
            all_courses[c['course_id']] = c
        else:
            dup_count += 1
    result = list(all_courses.values())
    print(f'  114上: {len(courses_shang)} 筆  |  114下: {len(courses_xia)} 筆')
    print(f'  略過重複: {dup_count} 筆')
    print(f'  合併後唯一課程數: {len(result)} 筆')

    with open(OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f'\n已儲存 → {OUTPUT}')

    print('\n前 5 筆範例:')
    for c in result[:5]:
        print(f"  {c['course_id']} | {c['course_name']} | "
              f"{c['credits']}學分 | {c['professor_name']} | "
              f"{c['department']} | {c['semester']}")


if __name__ == '__main__':
    if hasattr(sys.stdout, 'buffer') and sys.stdout.encoding.lower() in ('cp950', 'cp936', 'mbcs'):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    main()
